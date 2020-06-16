import urllib
import requests
from bs4 import BeautifulSoup
import re
import openpyxl
import time
import os
import random
import sys

#https://hackernoon.com/how-to-scrape-google-with-python-bo7d2tal
#clean company names by removing [" Corporation", ", Inc.", " Inc.", " Ltd.", " Incorporated"]
#worksheet headers: 1 Ticker, 2 Security Tickers, 3 $USD, 4 Company Name, 5 Glassdoor Reviews Link 6 GR Match % 7 Glassdoor Rating 8 Glassdoor Overview Link 9 GO Match % 10 Glassdoor Findings 11 Indeed Reviews Link 12 IR Match % 13 IR Finding

class Google_Scraper:
    def __init__(self, inputfile, outputfile):
        self.inputfile = inputfile
        self.outputfile = outputfile

    def load_xlsx(self):
        print(f"File {self.inputfile} opened")
        self.wb = openpyxl.load_workbook(self.inputfile, data_only=True) #parses the actual values and not the formulas
        # self.wb.create_sheet("Downloadlinks")
        self.ws = self.wb["Sheet1"] #current worksheet

        # for i in range(5): #create new headers for current worksheet
        #     self.ws.cell(column = i+1, row = 1).value = ["Firmnumber","ISIN","Date","Titel","Link"][i]

    def save_xlsx(self):
        self.wb.save(self.outputfile)
        print(f"File {self.outputfile} saved\n")

    def companies(self): #set company list which google_links will be looped through
        start, end = 1201, 2985 # script starts at row start and ends at row end ; in order to loop through all rows - > self.ws.max_row
        for col in self.ws.iter_cols(min_row=start, max_row=end, min_col=4, max_col=4): #iterate through company name column (3) -> only adjust max_row
            for cell in col:
                self.currentrow = cell.row
                self.firmname = cell.value #eg Amazon; isinstance(self.firmname, str) == True
                print(f"Current firm: {self.firmname} {self.currentrow} / {end}")
                googlesleep, indeedsleep = round(random.uniform(1.7,3.3), 2), round(random.uniform(8,12), 2) #sleep for 1 - 2.5 secs (float with 2 decimals, eg 2.26). 2,4 seems to be sufficient for bing. 20-40 is not sufficient for google, since bot prevention kicks in after around 80 queries
                
                print(f"Checking for Glassdoor links")
                print(f"Sleeping for {googlesleep}") 
                time.sleep(googlesleep)
                # self.google_links() #runs method to grab links from google
                self.bing_links() #runs method to grab links from bing

                # print(f"Checking for Indeed links")
                # print(f"Sleeping for {indeedsleep}") 
                # time.sleep(indeedsleep) #sleep for between 1 - 2.5 secs (float with 2 decimals, eg 2.26)
                # self.indeed_links() #runs method to grab indeed links from google

                self.save_xlsx() #save file after each firmname
            print("All done")

    def google_links(self): #method to grab google links for one company

        USER_AGENTS = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0", 
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.61 Safari/537.36",
                        # "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36",
                        # "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36",
                        # "Opera/9.80 (Windows NT 6.1; WOW64) Presto/2.12.388 Version/12.18"
                        ]

        headers = {"user-agent":random.choice(USER_AGENTS)}
        # print("Current header:", headers)

        query = f"site:glassdoor.com {self.firmname} reviews".replace(' ', '+')
        URL = f"https://google.com/search?q={query}&lr=lang_en&hl=en" #https://sites.google.com/site/tomihasa/google-language-codes

        url = requests.get(URL, headers=headers)
        soup = BeautifulSoup(url.content, "html.parser")

        global review_done, overview_done 
        review_done = 0
        overview_done = 0 #if both == 1 then break current loop and start next method

        if soup.find_all('div', class_='rc'):
            for g in soup.find_all('div', class_='rc'): #loop through every google search result (includes link, title, rating, body)
                link_and_title = g.find_all('a', limit=1) #get the link and title within current search result
                rating_and_reviews = g.find_all('div', class_="dhIWPd f") #get the div displaying the rating and amount of reviews
                if link_and_title: #if link and title is found
                    company_words = self.firmname.lower().split() #split the complete company name into separate strings
                    link = link_and_title[0]['href'] #get the href of the first list item
                    # title = g.find('h3').text #get the title

                    if re.search("glassdoor.com/Reviews/", link) and review_done != 1: #check if this part is in link
                        matching = 0 #split company string in separate words and check how many of them appear in the link
                        for i in company_words:
                            if re.search(i, link.lower()):
                                matching += 1
                        # print(title, link)
                        self.ws.cell(row = self.currentrow, column = 5).value = link
                        self.ws.cell(row = self.currentrow, column = 6).value = round(matching/len(company_words), 2)
                        # print(matching, "/", len(company_words), "matches in review link") #the higher the more likely the link will match the company
                        if rating_and_reviews: #if found review link also has a rating and amount of reviews
                            self.ws.cell(row = self.currentrow, column = 7).value = rating_and_reviews[0].text #eg Rating: 3,9 - ‎16 reviews
                        review_done = 1

                    if re.search("glassdoor.com/Overview/", link) and overview_done != 1: #check if this part is in link
                        matching = 0 #split company string in separate words and check how many of them appear in the link
                        for i in company_words:
                            if re.search(i, link.lower()):
                                matching += 1
                        # print(title, link)
                        self.ws.cell(row = self.currentrow, column = 8).value = link
                        self.ws.cell(row = self.currentrow, column = 9).value = round(matching/len(company_words), 2)
                        # print(matching, "/", len(company_words), "matches in overview link") #the higher the more likely the link will match the company
                        overview_done = 1

                    if review_done == 1 and overview_done == 1:
                        self.ws.cell(row = self.currentrow, column = 10).value = 3
                        print("Glassdoor: reviews and overview found")
                        break #break current loop and start next method/company

            if review_done == 1 and overview_done != 1:
                self.ws.cell(row = self.currentrow, column = 10).value = 1
                print("Glassdoor: only reviews found")
            elif overview_done == 1 and review_done != 1:
                self.ws.cell(row = self.currentrow, column = 10).value = 2
                print("Glassdoor: only overview found")
            elif review_done == 0 and overview_done == 0:
                self.ws.cell(row = self.currentrow, column = 10).value = 0
                print("Glassdoor: no links found on the first google page")
            
        else:
            print("NO LINKS FOUND - BOT PREVENTION ACTIVE")
            sys.exit()


    def bing_links(self): #method to grab bing links for one company

        USER_AGENTS = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0", 
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.61 Safari/537.36",
                        # "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36",
                        # "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36",
                        # "Opera/9.80 (Windows NT 6.1; WOW64) Presto/2.12.388 Version/12.18"
                        ]

        headers = {"user-agent":random.choice(USER_AGENTS)}
        # print("Current header:", headers)

        query = f"site:glassdoor.com {self.firmname} reviews".replace(' ', '+')
        URL = f"https://www.bing.com/search?q={query}"

        url = requests.get(URL, headers=headers)
        soup = BeautifulSoup(url.content, "html.parser")

        global review_done, overview_done 
        review_done = 0
        overview_done = 0 #if both == 1 then break current loop and start next method

        if soup.findAll("li", {"class":"b_algo"}):
            for g in soup.findAll("li", {"class":"b_algo"}): #loop through every google search result (includes link, title, rating, body)
                link_and_title = g.find_all('a', limit=1) #get the link and title within current search result
                # print(link_and_title)
                rating_and_reviews = g.find_all('div', class_="b_sritem") #get the div displaying the rating and amount of reviews
                if link_and_title: #if link and title is found
                    company_words = self.firmname.lower().split() #split the complete company name into separate strings
                    link = link_and_title[0]['href'] #get the href of the first list item
                    # title = g.find('h3').text #get the title

                    if re.search("glassdoor.com/Reviews/", link) and review_done != 1: #check if this part is in link
                        matching = 0 #split company string in separate words and check how many of them appear in the link
                        for i in company_words:
                            if re.search(i, link.lower()):
                                matching += 1
                        # print(title, link)
                        self.ws.cell(row = self.currentrow, column = 5).value = link
                        self.ws.cell(row = self.currentrow, column = 6).value = round(matching/len(company_words), 2)
                        # print(matching, "/", len(company_words), "matches in review link") #the higher the more likely the link will match the company
                        if rating_and_reviews: #if found review link also has a rating and amount of reviews
                            self.ws.cell(row = self.currentrow, column = 7).value = rating_and_reviews[0].text #eg Rating: 3,9 - ‎16 reviews
                        review_done = 1

                    if re.search("glassdoor.com/Overview/", link) and overview_done != 1: #check if this part is in link
                        matching = 0 #split company string in separate words and check how many of them appear in the link
                        for i in company_words:
                            if re.search(i, link.lower()):
                                matching += 1
                        # print(title, link)
                        self.ws.cell(row = self.currentrow, column = 8).value = link
                        self.ws.cell(row = self.currentrow, column = 9).value = round(matching/len(company_words), 2)
                        # print(matching, "/", len(company_words), "matches in overview link") #the higher the more likely the link will match the company
                        overview_done = 1

                    if review_done == 1 and overview_done == 1:
                        self.ws.cell(row = self.currentrow, column = 10).value = 3
                        print("Glassdoor: reviews and overview found")
                        break #break current loop and start next method/company

            if review_done == 1 and overview_done != 1:
                self.ws.cell(row = self.currentrow, column = 10).value = 1
                print("Glassdoor: only reviews found")
            elif overview_done == 1 and review_done != 1:
                self.ws.cell(row = self.currentrow, column = 10).value = 2
                print("Glassdoor: only overview found")
            elif review_done == 0 and overview_done == 0:
                self.ws.cell(row = self.currentrow, column = 10).value = 0
                print("Glassdoor: no links found on the first google page")

        else:
            print("NO LINKS FOUND - BOT PREVENTION ACTIVE?")
            # sys.exit()


    def indeed_links(self): #method to grab google links for one company

        USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0"
        headers = {"user-agent":USER_AGENT}

        query = f"indeed.com {self.firmname} reviews".replace(' ', '+')
        URL = f"https://google.com/search?q={query}&lr=lang_en&hl=en" #https://sites.google.com/site/tomihasa/google-language-codes

        url = requests.get(URL, headers=headers)
        soup = BeautifulSoup(url.content, "html.parser")

        global review_done, overview_done 
        review_done = 0 #
        overview_done = 0 #if both == 1 then break current loop and start next method

        for g in soup.find_all('div', class_='rc'): #loop through every google search result (includes link, title, rating, body)
            link_and_title = g.find_all('a', limit=1) #get the link and title within current search result
            if link_and_title: #if link and title is found
                company_words = self.firmname.lower().split() #split the complete company name into separate strings
                link = link_and_title[0]['href'] #get the href of the first list item
                # title = g.find('h3').text #get the title

                if re.search("indeed.com/cmp/", link) and review_done != 1: #check if this part is in link
                    matching = 0 #split company string in separate words and check how many of them appear in the link
                    for i in company_words:
                        if re.search(i, link.lower()):
                            matching += 1
                    # print(title, link)
                    self.ws.cell(row = self.currentrow, column = 11).value = link
                    self.ws.cell(row = self.currentrow, column = 12).value = round(matching/len(company_words), 2)
                    # print(matching, "/", len(company_words), "matches in review link") #the higher the more likely the link will match the company
                    review_done = 1
                    break

        if review_done == 1:
            self.ws.cell(row = self.currentrow, column = 13).value = 1
            print("Indeed: reviews found")
        else:
            self.ws.cell(row = self.currentrow, column = 13).value = 0
            print("Indeed: no links found on the first google page")


doc1 = Google_Scraper(inputfile = "C://Users//hannez//Desktop//Russell3000.xlsx", outputfile = "C://Users//hannez//Desktop//Russell3000.xlsx")
# doc1 = Google_Scraper(inputfile = "C://Users//hannez//Desktop//10.xlsx", outputfile = "C://Users//hannez//Desktop//10.xlsx")

#1 get links
doc1.load_xlsx()
doc1.companies()